import os
import io
import boto3
import textwrap

from openpyxl import load_workbook
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter

from django.conf import settings
from django.utils import timezone
from django.conf import settings

from dashboard.models import AddOnService
from dashboard.excel.service_calculator import ServiceSheetCalculator, to_nengo, format_comma

import logging
logger = logging.getLogger(__name__)

def create_service_sheet(context):
    try:
        wb = load_workbook('templatesExcel/service_template.xlsx')
        calc = ServiceSheetCalculator(context)
        res = calc.get_results()
        
        user = context['user']
        year, month = context['dis_year'], context['dis_month']
        ws1 = wb['1'] # スケジュール

        # --- Sheet 1: ヘッダー (生活保護情報の印字) ---
        if res['is_hogo']:
            pa = res['pa_data']
            ws1['G5'] = "25" # 法別番号
            ws1['I5'] = pa.hogo_number # 保護番号(8桁)
            ws1['V9'] = pa.recipient_number # 受給者番号(10桁)
        
        # (中略: スケジュールのループ書き込み)

        # --- Sheet 2: 別表（実績） ---
        ws2 = wb['2']
        current_row = 6
        office_name = context['office'].name.split()[:2]

        # 明細書き込み
        for item in res['plan_items'] + res['addon_items']:
            write_billing_line(ws2, current_row, item, context['office'].office_number, office_name)
            current_row += 1

        # 小計行 (地域密着通所合計)
        ws2[f'AT{current_row}'] = format_comma(res['seikyu_taisyu']) # 10割分
        ws2[f'AX{current_row}'] = format_comma(res['insurance_seikyu']) # 保険分(9割)
        
        # 本人負担欄 (BD列) の制御
        if res['is_hogo']:
            ws2[f'BA{current_row}'] = format_comma(res['public_seikyu']) # 公費分(1割)
            ws2[f'BD{current_row}'] = "0" # 本人は0
        else:
            ws2[f'BD{current_row}'] = format_comma(res['user_hutan']) # 本人が払う

        # 最終合計行 (20行目固定)
        ws2['AT20'] = format_comma(res['seikyu_taisyu'])
        ws2['AX20'] = format_comma(res['insurance_seikyu'])
        if res['is_hogo']:
            ws2['BA20'] = format_comma(res['public_seikyu'])
            ws2['BD20'] = "0"
        else:
            ws2['BD20'] = format_comma(res['user_hutan'])

        # --- 生活保護情報の印字を追加 ---
        bikou_text =''
        if res['is_hogo']:
            pa = res['pa_data']
            # セル位置はテンプレートに合わせて調整してください
            bikou_text += "生活保護の法別番号:25\n" # 生活保護の法別番号
            bikou_text += f'公費負担者番号:{pa.hogo_number}\n'      # 公費負担者番号
            bikou_text += f'公費受給者番号:{pa.recipient_number}' # 公費受給者番号
        ws2['AK23'] = bikou_text
# ローカル保存処理
        if settings.DJANGO_ENV == 'dev':
            filepath, filename = get_service_sheet_path(user, year, month)
            wb.save(filepath)

# s3の場合保存処理
        else:
            buffer = io.BytesIO()
            wb.save(buffer)
            buffer.seek(0)
            file_bytes = buffer.getvalue()
            key, filename = get_service_sheet_path(user, year, month)
            upload_service_sheet_to_s3(key,file_bytes)
        
# 後処理
        _record_model_update(user, year, month,res)
        return
    except Exception as e:
        logger.error(f"サービス提供表の作成中にエラーが発生しました: {e}")
        raise

def write_billing_line(ws, row, item, office_number, office_names):
    """別表の1行分を書き込むヘルパー"""
    ws[f'A{row}'] = "\n".join(office_names)
    ws[f'A{row}'].alignment = Alignment(wrap_text=True)
    ws[f'G{row}'] = office_number
    _auto_newline(item['name'], ws, f'L{row}', 6)
    ws[f'Q{row}'] = item['code']
    ws[f'T{row}'] = item['unit']
    ws[f'Y{row}'] = item['count']
    ws[f'Z{row}'] = format_comma(item['subtotal'])
    ws[f'AC{row}'] = format_comma(item['subtotal'])

def _auto_newline(text, ws, cell, line=10):
    '''セル内で自動改行するためのヘルパー関数'''
    text = str(text)
    wrapped = "\n".join(textwrap.wrap(text, line))
    ws[cell] = wrapped
    ws[cell].alignment = Alignment(wrap_text=True, vertical="center",)

def _record_model_update(user, year, month, res):
    """ServiceMonthlyRecordに計算結果を反映する"""
    from dashboard.models import ServiceMonthlyRecord
    target_date = timezone.datetime(year, month, 1).date()
    
    record, _ = ServiceMonthlyRecord.objects.get_or_create(
        user=user, 
        date=target_date
    )
    
    # 計算結果をモデルのフィールドに保存
    record.total_cost = res['seikyu_taisyu']
    record.benefit_amount = res['insurance_seikyu'] # 保険請求額
    record.public_amount = res['public_seikyu']  # 公費請求額 
    record.user_share_amount = res['user_hutan']  # 利用者負担額
    record.within_units = res['within_units']
    record.over_units = res['over_units']
    
    record.confirmed = True
    record.confirmed_at = timezone.now().date()
    record.save()
    logger.info(f"{user.name}様の{year}年{month}月分実績を保存しました")

def get_service_sheet_path(user, year, month):
    '''サービス提供表の保存パスを取得するユーティリティ関数。ユーザーIDと年月を指定して、ローカルまたはS3のパスを返す。'''
    if settings.DJANGO_ENV == 'dev':
        user_dir = os.path.join(
            settings.MEDIA_ROOT,
            "service_sheets_export",
            f"{user.id}_{user.name}"
        )
        os.makedirs(user_dir, exist_ok=True)
        year_month_dir = f"{year}_{month:02d}"
        date_dir = os.path.join(user_dir, year_month_dir)
        os.makedirs(date_dir, exist_ok=True)
        filename = f"サービス提供表_{user.name}_{year}_{month}.xlsx"
        return os.path.join(date_dir, filename), filename
# s3バージョン
    else:
        key = f"service_sheets_export/{user.id}_{user.name}/{year}_{month:02d}/サービス提供表_{user.name}_{year}_{month}.xlsx"
        filename = f"サービス提供表_{user.name}_{year}_{month}.xlsx"
        return key, filename

def upload_service_sheet_to_s3(key, file_bytes):
    s3 = boto3.client('s3')
    s3.put_object(
        Bucket=settings.AWS_STORAGE_BUCKET_NAME,
        Key=key,
        Body=file_bytes,
        ContentType='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    return f"https://{settings.AWS_STORAGE_BUCKET_NAME}.s3.amazonaws.com/{key}"
def add_comma(value):
    return f"{int(value):,}"