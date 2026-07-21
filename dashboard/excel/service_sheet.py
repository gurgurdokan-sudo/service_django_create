import os
import io
import boto3
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
from dashboard.models import User, Office, Municipality, AddOnService
from dashboard.calendar_table import get_month_days
from dashboard.excel.service_salculator import ServiceSheetCalculator,to_nengo, format_comma
from django.utils import timezone
from django.conf import settings
from django.http import FileResponse
import textwrap
    
import logging
logger = logging.getLogger(__name__)

def create_service_sheet(context):
    try:
        wb = load_workbook('templatesExcel/service_template.xlsx')
        calc = ServiceSheetCalculator(context)
        res = calc.get_results()
        
        office = context['office']
        user = context['user']
        year, month = context['dis_year'], context['dis_month']
        now = timezone.now()
        office_name = (office.name.split() + ["", ""])[:2]

        # --- Sheet 1: スケジュール ---
        ws = wb['1']
        ws.title = 'スケジュール'
        ws['B2'] = '認定済'

        # 市町村コード・名称
        for i, d in enumerate(office.municipality.municipality_code):
            ws[f'{get_column_letter(11+i)}5'] = d
        ws['V5'] = office.municipality.name

        # ケアマネ情報
        cm = user.care_manager
        ws['Aj5'] = cm.office_name if cm else ''
        ws['Aj6'] = cm.name if cm else ''

        # 作成日・届出日
        ws['AX5'] = f"{to_nengo(now.year, now.month)} {now.month}月 {now.day}日"
        ws['AX7'] = f"{to_nengo(year, month)} {month}月 1日"

        # 被保険者情報
        for i, d in enumerate(str(user.insured_number)):
            ws[f"{get_column_letter(7+i)}7"] = d
        ws['V7'], ws['V8'] = user.name_kana, user.name
        
        # 生年月日・性別
        birth = user.date_of_birth
        ws['G9'] = to_nengo(birth.year, birth.month)
        ws['G11'] = f"{birth.month}月{birth.day}日"
        ws['N9'] = user.get_gender_display()[:1]

        # 介護度・限度額
        ws['W9'] = user.care_level
        if user.old_certificate:
            ws['W11'] = user.old_certificate.care_level
            date = user.latest_changed_date
            ws['W12'] = f"{to_nengo(date.year, date.month)} {date.month}月 {date.day}日"
        ws['AI9'] = user.max_separate_payment
        ws['BE9'] = '0'

        # --- ★カレンダーヘッダー (15行目・16行目) ---
        calendar = context['calendar']
        for i, d in enumerate(calendar):
            col = get_column_letter(25 + i)
            ws[f'{col}15'] = d['day']
            ws[f'{col}16'] = '◎' if d['is_holiday'] else d['weekday_jp']

        # サービス行の書き込み (元のループ構造を維持)
        row = 17
        for plan in context['plans']:
            ws[f'B{row}'] = f'{plan.start_time:%H:%M}～{plan.end_time:%H:%M}'
            _auto_newline(plan.service_name, ws, f'H{row}')
            ws[f'O{row}'], ws[f'O{row+1}'] = office_name[0], office_name[1]
            for i, c in enumerate(calendar):
                col = get_column_letter(25 + i)
                ws[f'{col}{row}'] = plan.schedule_dict.get(str(c['day']), '')
                ws[f'{col}{row+1}'] = plan.actual_dict.get(str(c['day']), {}).get('main', '')
            ws[f'BD{row}'] = plan.get_total_count('schedule')
            ws[f'BD{row+1}'] = plan.get_total_count('actual')
            row += 2
        # Addonの自費は除外
        self_pay_usage_map = {}
        self_pay_addon_names = AddOnService.objects.filter(
            insurance_type=self_pay).values_list('service_name', flat=True)
        # 加算行の書き込み (plan.get_addon_summary を使用)
        for plan in context['plans']:
            addon_summary = plan.get_addon_summary
            if not addon_summary: continue
            for addon_name, days_list in addon_summary.items():
                if addon_name in self_pay_addon_names:
                    self_pay_usage_map[addon_name] = days_list #todo 後で自費
                    continue
                _auto_newline(addon_name, ws, f'H{row}')
                ws[f'O{row}'], ws[f'O{row+1}'] = office_name[0], office_name[1]
                for i, c in enumerate(calendar):
                    col = get_column_letter(25 + i)
                    day = str(c['day'])
                    ws[f'{col}{row}'] = plan.schedule_dict.get(day, '')
                    ws[f'{col}{row+1}'] = "1" if day in days_list else ""
                ws[f'BD{row}'] = plan.get_total_count('schedule')
                ws[f'BD{row+1}'] = len(days_list)
                row += 2

        # デフォルトサービス（特定事業所加算等）
        if office.default_service:
            _auto_newline(office.default_service.service_name, ws, f'H{row}')
            ws[f'O{row}'], ws[f'O{row+1}'] = office_name[0], office_name[1]
            ws[f'BD{row}'], ws[f'BD{row+1}'] = '1', '1'
            row += 2

        # --- Sheet 2: 別表（実績） ---
        ws = wb['2']
        ws.title = '別表（実績）'
        # (ヘッダー情報...)
        
        current_row = 6
        # プラン明細
        for item in res['plan_items']:
            write_billing_line(ws, current_row, item, office.office_number, office_name)
            current_row += 1
        # 加算明細
        for item in res['addon_items']:
            write_billing_line(ws, current_row, item, office.office_number, office_name)
            current_row += 1

        # 小計（地域密着型通所合計）
        kako = ['(', ')']
        ws[f'A{current_row}'] = '\n'.join(office_name)
        _auto_newline('地域密着通所合計', ws, f'L{current_row}', 6)
        ws[f'Z{current_row}'] = f"({add_comma(res['subtotal_units'])})"
        ws[f'AC{current_row}'] = f"({add_comma(res['subtotal_units'])})"
        ws[f'AO{current_row}'] = add_comma(res['subtotal_units'])
        ws[f'AR{current_row}'] = calc.unit_price
        ws[f'AT{current_row}'] = add_comma(res['seikyu_taisyu'])
        ws[f'AW{current_row}'] = int(calc.benefit_rate * 100)
        ws[f'AX{current_row}'] = add_comma(res['seikyu_bun'])
        ws[f'BD{current_row}'] = add_comma(res['hutan'])

        # デフォルト行
        if office.default_service:
            current_row += 1
            ws[f'A{current_row}'] = '\n'.join(office_name)
            ws[f'G{current_row}'] = office_number
            _auto_newline(office.default_service.service_name, ws, f'L{current_row}', 6)
            ws[f'Z{current_row}'] = f"({add_comma(res['def_unit'])})"
            ws[f'AO{current_row}'] = f"({add_comma(res['def_unit'])})"
            ws[f'AR{current_row}'] = calc.unit_price
            ws[f'AT{current_row}'] = add_comma(res['def_total_cost'])
            ws[f'AW{current_row}'] = int(calc.benefit_rate * 100)
            ws[f'AX{current_row}'] = add_comma(res['def_benefit'])
            ws[f'BD{current_row}'] = add_comma(res['def_user_share'])

        # 最終合計 (row 20 固定)
        ws['T20'] = add_comma(user.max_separate_payment)
        ws['Z20'] = add_comma(res['subtotal_units'])
        ws['AC20'] = add_comma(res['subtotal_units'])
        ws['AL20'] = add_comma(res['over_units']) if res['over_units'] > 0 else ''
        ws['AO20'] = add_comma(res['within_units'])
        ws['AT20'] = add_comma(res['total_taisyou'])
        ws['AX20'] = add_comma(res['total_seikyu'])
        ws['BD20'] = add_comma(res['total_hutan'])
        ws['BG20'] = add_comma(res['over_full_share']) if res['over_full_share'] > 0 else ''
# ローカル保存処理
        # filepath, filename = get_service_sheet_path(user, year, month)
        # wb.save(filepath)

# s3の場合保存処理
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        file_bytes = buffer.getvalue()
        key, filename = get_service_sheet_path(user, year, month)
        upload_service_sheet_to_s3(key,file_bytes)
        
        _recode_model_create(user, year, month)
        return
    except Exception as e:
        logger.error(f"サービス提供表の作成中にエラーが発生しました: {e}")
        raise

def write_billing_line(ws, row, item, office_number, office_names):
    """別表の1行分を書き込むヘルパー"""
    ws[f'A{row}'] = "\n".join(office_names)
    ws[f'A{row}'].alignment = Alignment(wrap_text=True)
    ws[f'G{row}'] = office_number
    _auto_newline(item['name'], ws, f'L{row}', 6)
    ws[f'Q{row}'] = item['code']
    ws[f'T{row}'] = item['unit']
    ws[f'Y{row}'] = item['count']
    ws[f'Z{row}'] = format_comma(item['subtotal'])
    ws[f'AC{row}'] = format_comma(item['subtotal'])

def _auto_newline(text, ws, cell, line=10):
    '''セル内で自動改行するためのヘルパー関数'''
    text = str(text)
    wrapped = "\n".join(textwrap.wrap(text, line))
    ws[cell] = wrapped
    ws[cell].alignment = Alignment(wrap_text=True, vertical="center",)

def _recode_model_create(user, year, month):
    '''ServiceMonthlyRecordモデルにレコードを作成する'''
    from dashboard.models import ServiceMonthlyRecord
    date = timezone.datetime(year, month, 1)
    record = ServiceMonthlyRecord.objects.filter(user=user, date=date).first()
    if not record:
        record = ServiceMonthlyRecord(
            user = user,
            date = date,
            path = get_service_sheet_path(user, year, month)[0],
            total_seikyu = res['total_seikyu'],
            addon_cords = res['addon_items']['cord'],
            total_hutan = res['total_hutan'],
            within_units = res['within_units'],
        )
    record.confirmed = True
    record.save()
def get_service_sheet_path(user, year, month):#Pathを返す
    # user_dir = os.path.join(
    #     settings.MEDIA_ROOT,
    #     "service_sheets_export",
    #     f"{user.id}_{user.name}"
    # )
    # os.makedirs(user_dir, exist_ok=True)
    # year_month_dir = f"{year}_{month:02d}"
    # date_dir = os.path.join(user_dir, year_month_dir)
    # os.makedirs(date_dir, exist_ok=True)
    # filename = f"サービス提供表_{user.name}_{year}_{month}.xlsx"
    # return os.path.join(date_dir, filename), filename
# s3バージョン
    key = f"service_sheets_export/{user.id}_{user.name}/{year}_{month:02d}/サービス提供表_{user.name}_{year}_{month}.xlsx"
    filename = f"サービス提供表_{user.name}_{year}_{month}.xlsx"
    return key, filename

def upload_service_sheet_to_s3(key, file_bytes):
    s3 = boto3.client('s3')
    s3.put_object(
        Bucket=settings.AWS_STORAGE_BUCKET_NAME,
        Key=key,
        Body=file_bytes,
        ContentType='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    return f"https://{settings.AWS_STORAGE_BUCKET_NAME}.s3.amazonaws.com/{key}"
def add_comma(value):
    return f"{int(value):,}"